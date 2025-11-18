import streamlit as st
import pandas as pd
import numpy as np
try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except Exception as e:
    # Não interrompe o app — registramos e exibimos uma mensagem ao usuário onde fizer sentido
    logging.warning(f"Plotly import falhou: {e}")
    go = None
    PLOTLY_AVAILABLE = False
from datetime import timedelta
from typing import List, Tuple
import inspect
import warnings
import logging
import os
from sklearn.metrics import mean_absolute_error, mean_squared_error
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill

# Suprimir avisos de deprecação do Plotly e bibliotecas relacionadas
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', message='.*keyword arguments.*deprecated.*')
warnings.filterwarnings('ignore', message='.*use_container_width.*')
warnings.simplefilter('ignore')

# Configurar logs
os.makedirs(os.path.join(os.getcwd(), 'logs'), exist_ok=True)
LOGFILE = os.path.join(os.getcwd(), 'logs', 'app.log')
logging.basicConfig(filename=LOGFILE, level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')

# Importar modelos de serie temporal
try:
    from prophet import Prophet
except Exception as e:
    logging.warning(f"Prophet import falhou: {e}")
    Prophet = None

from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.tsa.holtwinters import ExponentialSmoothing

logging.getLogger('prophet').setLevel(logging.WARNING)

# ==================== FUNCOES DE METRICAS ====================

def calcular_metricas(y_real, y_pred, nome_modelo=''):
    """Calcula MAE, RMSE e MAPE para avaliar modelos."""
    mae = mean_absolute_error(y_real, y_pred)
    rmse = np.sqrt(mean_squared_error(y_real, y_pred))
    # MAPE (Mean Absolute Percentage Error)
    mape = np.mean(np.abs((y_real - y_pred) / y_real)) * 100 if (y_real != 0).all() else np.nan
    return {'MAE': mae, 'RMSE': rmse, 'MAPE': mape}

def validar_modelo_com_holdout(df, modelo_func, horizonte=7, test_size=0.2):
    """Valida modelo usando holdout (20% dos dados como teste)."""
    try:
        n_test = max(1, int(len(df) * test_size))
        df_train = df.iloc[:-n_test].copy()
        df_test = df.iloc[-n_test:].copy()
        
        if len(df_train) < 10:
            return None
        
        # Treina no conjunto de treino
        previ, erro = modelo_func(df_train, n_test)
        if erro or previ is None:
            return None
        
        # Extrai valores preditos
        col_p = 'yhat' if 'yhat' in previ.columns else 'previsao'
        y_pred = previ[col_p].values[:n_test]
        y_real = df_test['demanda'].values
        
        # Alinha tamanhos (em caso de diferenca)
        if len(y_pred) > len(y_real):
            y_pred = y_pred[:len(y_real)]
        elif len(y_real) > len(y_pred):
            y_real = y_real[:len(y_pred)]
        
        if len(y_real) == 0 or len(y_pred) == 0:
            return None
        
        return calcular_metricas(y_real, y_pred)
    except Exception as e:
        logging.warning(f"Erro ao validar modelo {modelo_func.__name__}: {e}")
        return None

def criar_excel_com_resultados(df_previsoes, df_metricas, local, sub_local, local_terciario):
    """Cria um arquivo Excel com abas de previsoes e metricas."""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Aba 1: Previsoes consolidadas
        df_previsoes.to_excel(writer, sheet_name='Previsoes', index=False)
        
        # Aba 2: Metricas
        df_metricas.to_excel(writer, sheet_name='Metricas', index=False)
        
        # Aba 3: Info
        info_data = {
            'Campo': ['Local', 'Nivel1', 'Nivel2', 'Data Geracao', 'Quantidade Modelos'],
            'Valor': [local, sub_local, local_terciario, pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'), len(df_metricas)]
        }
        df_info = pd.DataFrame(info_data)
        df_info.to_excel(writer, sheet_name='Info', index=False)
        
        # Formatar abas
        workbook = writer.book
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    output.seek(0)
    return output

def criar_nome_arquivo_excel(local, sub_local):
    """Cria nome limpo para arquivo Excel."""
    import re
    s = f"previsoes_{local}_{sub_local}"
    s = re.sub(r'[^\w\s-]', '', s).strip()
    s = re.sub(r'[\s-]+', '_', s)
    return f"{s}.xlsx"

# Configuracao da pagina
st.set_page_config(page_title="Previsao de Series Temporais", layout="wide")

# Titulo
st.title('Sistema de apoio a decisão - Previsão de demanda')
st.write('Plataforma integrada para previsao de demanda com modelos: Prophet (melhorado), ARIMA, SARIMA e Holt-Winters. Comparacoes e metricas de assertividade incluidas.')

# ==================== FUNCOES ====================

@st.cache_data
def carregar_dados(arquivo):
    """Carrega CSV ou Excel a partir do upload e normaliza colunas esperadas.

    Aceita um caminho ou um file-like (Streamlit UploadedFile). Detecta automaticamente
    se é Excel (.xlsx) ou CSV e tenta mapear colunas comuns para o formato esperado
    pelo app: `data`, `demanda`, `local`, `sub_local`, `local_terciario`.
    """
    try:
        filename = getattr(arquivo, 'name', None)
        if filename and str(filename).lower().endswith(('.xls', '.xlsx')):
            df = pd.read_excel(arquivo)
        else:
            df = pd.read_csv(arquivo)

        # Normalizar nomes de colunas
        df.columns = [c.strip() for c in df.columns]

        # Detectar colunas de data, local e demanda automaticamente
        # detectar colunas que contem data ou timestamp (aceitar 'timestamp' ou 'datetime')
        col_date = next((c for c in df.columns if any(k in c.lower() for k in ['data', 'date', 'timestamp', 'datetime', 'ts', 'data_hora', 'date_time'])), None)
        col_local = next((c for c in df.columns if any(k in c.lower() for k in ['rest', 'local', 'nome', 'restaurant', 'filial'])), None)
        col_demanda = next((c for c in df.columns if any(k in c.lower() for k in ['mesa', 'table', 'quant', 'qtd', 'demanda', 'count'])), None)
        # detectar hora (opcinal)
        # detectar coluna hora apenas por nomes isolados (ex.: 'hora', 'hour', 'time'), evitar match em 'datetime'
        import re
        col_hora = next((c for c in df.columns if re.search(r"\b(hora|hour|time|hh)\b", c.lower())), None)

        # Se não detectar todas, retorna df e colunas detectadas para etapa manual
        return df, col_date, col_local, col_demanda, col_hora
    except Exception as e:
        logging.exception(f"Erro ao carregar arquivo: {e}")
        raise

def validar_dados(df):
    cols_necessarias = ['data', 'demanda', 'local', 'sub_local', 'local_terciario']
    cols_faltando = [c for c in cols_necessarias if c.lower() not in [x.lower() for x in df.columns]]
    
    if cols_faltando:
        return False, f"Colunas faltando: {', '.join(cols_faltando)}"
    
    try:
        df.columns = [c.lower().strip() for c in df.columns]
        df['data'] = pd.to_datetime(df['data'])
        df['demanda'] = pd.to_numeric(df['demanda'])
        return True, df
    except Exception as e:
        logging.exception(f"Erro na validacao de dados: {e}")
        return False, f"Erro: {str(e)}"


def _assign_shift(ts: pd.Timestamp, shifts: List[Tuple[int, int]]):
    """Assign a shift index to a timestamp given a list of (start_hour, end_hour).
       If shift wraps midnight (start > end) assign accordingly.
       Returns index (int) or None if not in any shift."""
    if pd.isna(ts):
        return None
    hour = int(ts.hour)
    date = ts.normalize()
    for i, (start, end) in enumerate(shifts):
        if start <= end:
            if start <= hour <= end:
                return i
        else:
            # wrap midnight
            if hour >= start or hour <= end:
                return i
    return None


def aggregate_shifts(df: pd.DataFrame, shifts: List[Tuple[int, int]], agg: str = 'Soma') -> pd.DataFrame:
    """Aggregate hourly data into shift-level records.
       Returns DataFrame with columns: data (timestamp at shift start), demanda, local, sub_local, local_terciario, turno
    """
    try:
        # work on a copy
        df_h = df.copy()
        # assign shift index
        df_h['turno_idx'] = df_h['data'].apply(lambda d: _assign_shift(d, shifts))
        # ignore rows not in any shift
        df_h = df_h[df_h['turno_idx'].notna()].copy()
        if len(df_h) == 0:
            logging.error("Nenhuma hora encontrada nos turnos definidos")
            return None
        df_h['turno_idx'] = df_h['turno_idx'].astype(int)

        # determine shift day: for wrapped shifts, if hour <= end then -1 day so shift belongs to previous day
        def compute_shift_day(r):
            try:
                idx = int(r['turno_idx'])
                if idx < 0 or idx >= len(shifts):
                    return None
                start, end = shifts[idx]
                dt = r['data']
                hr = dt.hour
                if start <= end:
                    # same-day shift: use date of dt
                    return pd.Timestamp(dt.date())
                else:
                    # wrapped shift: if hour <= end -> belongs to previous day (start on previous day)
                    if hr <= end:
                        return pd.Timestamp((dt - pd.Timedelta(days=1)).date())
                    return pd.Timestamp(dt.date())
            except Exception as e:
                logging.warning(f"Erro em compute_shift_day para linha {r.name}: {e}")
                return None

        # aplicar funcao linha por linha, conversao segura
        shift_dates = []
        for idx, row in df_h.iterrows():
            sd = compute_shift_day(row)
            shift_dates.append(sd)
        
        df_h['shift_date'] = pd.Series(shift_dates, index=df_h.index)
        df_h = df_h[df_h['shift_date'].notna()].copy()

        # aggregate
        if agg == 'Soma':
            df_agg = df_h.groupby(['shift_date', 'turno_idx', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].sum()
        elif agg == 'Média':
            df_agg = df_h.groupby(['shift_date', 'turno_idx', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].mean()
        elif agg == 'Máximo':
            df_agg = df_h.groupby(['shift_date', 'turno_idx', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].max()
        elif agg == 'Mínimo':
            df_agg = df_h.groupby(['shift_date', 'turno_idx', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].min()
        elif agg == 'Percentil 80%':
            # quantile retorna um index diferente, precisa de tratamento especial
            grouped = df_h.groupby(['shift_date', 'turno_idx', 'local', 'sub_local', 'local_terciario'])
            df_agg = grouped['demanda'].quantile(0.8).reset_index()
            df_agg.rename(columns={'demanda': 'demanda'}, inplace=True)
        else:
            df_agg = df_h.groupby(['shift_date', 'turno_idx', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].max()

        # create shift start datetime for 'data' (date + shift start hour)
        def make_shift_ts(row):
            s_hour = shifts[int(row['turno_idx'])][0]
            return pd.Timestamp(row['shift_date']) + pd.Timedelta(hours=int(s_hour))

        df_agg['data'] = df_agg.apply(make_shift_ts, axis=1)
        df_agg = df_agg.rename(columns={'demanda': 'demanda'})
        # sort and return
        df_agg = df_agg[['data', 'demanda', 'local', 'sub_local', 'local_terciario', 'turno_idx']].sort_values('data')
        # label turno
        df_agg['turno'] = df_agg['turno_idx'].apply(lambda x: f"Turno {x+1}")
        return df_agg
    except Exception as e:
        logging.error(f"Erro ao agregar turnos: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return None

def preparar_dados(df, local, sub_local, local_terciario):
    # Se o usuario selecionar 'Todos' em qualquer nivel, nao filtra por esse nivel
    mask = pd.Series(True, index=df.index)
    if local is not None and str(local).lower() != 'todos':
        mask &= (df['local'] == local)
    if sub_local is not None and str(sub_local).lower() != 'todos':
        mask &= (df['sub_local'] == sub_local)
    if local_terciario is not None and str(local_terciario).lower() != 'todos':
        mask &= (df['local_terciario'] == local_terciario)

    dados = df[mask].copy()
    return dados.sort_values('data')

# ==================== MODELOS ====================

def prever_prophet(df, horizonte=30, freq='D', future_custom: pd.DatetimeIndex = None):
    if Prophet is None:
        return None, "Prophet nao instalado"
    try:
        # Deteccao automatica de sazonalidade
        n_dados = len(df)
        # Ajuste de sazonalidade com base na frequência dos dados
        if freq == 'H':
            has_weekly = n_dados >= (7 * 24)
            has_yearly = n_dados >= (365 * 24)
        else:
            has_weekly = n_dados >= 14
            has_yearly = n_dados >= 365
        
        # Normalizar dados
        y_mean = df['demanda'].mean()
        y_std = df['demanda'].std()
        y_min = df['demanda'].min()
        
        df_prophet = df[['data', 'demanda']].rename(columns={'data': 'ds', 'demanda': 'y'}).copy()
        
        # Parametrizacao adaptativa baseada no tamanho dos dados
        changepoint_prior = 0.001 if n_dados < 100 else 0.01
        seasonality_prior = 5.0 if n_dados < 365 else 10.0
        
        # Build kwargs and ensure compatibility with installed Prophet signature
        kwargs = dict(
            yearly_seasonality=has_yearly,
            weekly_seasonality=has_weekly,
            daily_seasonality=True if freq == 'H' else False,
            interval_width=0.80,
            changepoint_prior_scale=changepoint_prior,
            seasonality_prior_scale=seasonality_prior,
            seasonality_mode='additive',
            growth='linear'
        )

        # Filter kwargs to only those accepted by the installed Prophet version
        try:
            sig = inspect.signature(Prophet.__init__)
            allowed = {p for p in sig.parameters.keys() if p not in ('self', 'args', 'kwargs')}
            kwargs = {k: v for k, v in kwargs.items() if k in allowed}
        except Exception:
            # If anything goes wrong with introspection, proceed without filtering
            pass

        try:
            model = Prophet(**kwargs)
        except TypeError as e:
            logging.warning(f"Prophet init failed with kwargs: {e}. Trying without optional kwargs.")
            # Try a minimal init as fallback (some Prophet versions have different param sets)
            try:
                model = Prophet()
            except Exception as e2:
                # re-raise original error for clarity
                logging.exception(f"Fallback Prophet init also failed: {e2}")
                raise
            
        
        model.fit(df_prophet)
        # O Prophet aceita argumento `freq` para gerar o futuro em horas/dias
        if future_custom is not None:
            future = pd.DataFrame({'ds': future_custom})
        else:
            future = model.make_future_dataframe(periods=horizonte, freq=freq)
        forecast = model.predict(future)
        
        # Aplicar constrains realistas
        # Demanda nao deve ser menor que a media menos 2 desvios
        lower_bound = max(0, y_min - y_std)
        forecast['yhat'] = forecast['yhat'].clip(lower=lower_bound)
        forecast['yhat_lower'] = forecast['yhat_lower'].clip(lower=0)
        
        return forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(horizonte), None
    except Exception as e:
        logging.exception(f"Erro Prophet: {e}")
        return None, str(e)

def _executar_modelo_statsmodels(df, modelo_func, nome_modelo, horizonte=30):
    """Função genérica para executar modelos do statsmodels."""
    try:
        modelo = modelo_func(df['demanda'])
        fitted = modelo.fit()
        
        # Holt-Winters tem um método de previsão diferente
        if isinstance(modelo, ExponentialSmoothing):
            previ = fitted.forecast(steps=horizonte)
            df_pre = pd.DataFrame({'previsao': previ.values})
            # Garantir previsões não-negativas
            df_pre['previsao'] = pd.to_numeric(df_pre['previsao'], errors='coerce').clip(lower=0)
            return df_pre, None
        
        # Lógica para ARIMA/SARIMAX
        forecast = fitted.get_forecast(steps=horizonte)
        previ = forecast.predicted_mean
        conf = forecast.conf_int(alpha=0.05)
        df_pre = pd.DataFrame({
            'previsao': previ.values,
            'inf': conf.iloc[:, 0].values,
            'sup': conf.iloc[:, 1].values
        })
        # Coerce numeric and remover limites negativos: a demanda nao pode ser negativa
        for c in ['previsao', 'inf', 'sup']:
            if c in df_pre.columns:
                df_pre[c] = pd.to_numeric(df_pre[c], errors='coerce')

        # Limitar lower bound a 0
        if 'previsao' in df_pre.columns:
            df_pre['previsao'] = df_pre['previsao'].clip(lower=0)
        if 'inf' in df_pre.columns:
            df_pre['inf'] = df_pre['inf'].clip(lower=0)

        # Garantir consistência: inf <= previsao <= sup
        if 'inf' in df_pre.columns and 'previsao' in df_pre.columns:
            df_pre.loc[df_pre['inf'] > df_pre['previsao'], 'inf'] = df_pre.loc[df_pre['inf'] > df_pre['previsao'], 'previsao']
        if 'sup' in df_pre.columns and 'previsao' in df_pre.columns:
            df_pre.loc[df_pre['sup'] < df_pre['previsao'], 'sup'] = df_pre.loc[df_pre['sup'] < df_pre['previsao'], 'previsao']

        return df_pre, None
    except Exception as e:
        logging.exception(f"Erro {nome_modelo}: {e}")
        return None, str(e)

def prever_arima(df, horizonte=30):
    return _executar_modelo_statsmodels(df, lambda x: ARIMA(x, order=(1, 1, 1)), "ARIMA", horizonte)

def prever_sarima(df, horizonte=30, seasonal_periods=None):
    s = 12 if seasonal_periods is None else seasonal_periods
    # Ensure seasonal period > 1 for SARIMAX; if not, disable seasonal component
    if s is None or int(s) <= 1:
        seasonal_order = (0, 0, 0, 0)
        logging.info(f"SARIMA: sazonalidade desativada por periodicidade s={s}")
    else:
        seasonal_order = (1, 1, 1, int(s))
    return _executar_modelo_statsmodels(df, lambda x: SARIMAX(x, order=(1, 1, 1), seasonal_order=seasonal_order), "SARIMA", horizonte)

def prever_holt_winters(df, horizonte=30, seasonal_periods=None):
    params = {'trend': 'add'}
    # Ajuste automatico: se houver dados suficientes, usa sazonalidade com seasonal_periods
    sp = seasonal_periods if seasonal_periods is not None else 12
    # Only enable seasonal when seasonal_periods > 1 and there is sufficient data
    if sp is not None and int(sp) > 1 and len(df) >= max(26, sp * 2):
        params.update({'seasonal': 'add', 'seasonal_periods': int(sp)})
    else:
        logging.info(f"Holt-Winters: sazonalidade nao aplicada (sp={sp}, len(df)={len(df)})")
    return _executar_modelo_statsmodels(df, lambda x: ExponentialSmoothing(x, **params), "Holt-Winters", horizonte)

# ==================== APP ====================

uploaded_file = st.file_uploader("Upload CSV ou Excel (.csv / .xlsx) com series temporais", type=["csv", "xlsx"]) 
try:
    if uploaded_file is not None:
        df_raw, col_date, col_local, col_demanda, col_hora = carregar_dados(uploaded_file)
        # Se não detectar todas as colunas, pedir para o usuário selecionar manualmente
        # Mapeamento manual ou automático
        # Sempre obrigar mapear as colunas inicialmente (persistente por upload)
        file_key = getattr(uploaded_file, 'name', None)
        # reset mapping when file changes
        if st.session_state.get('mapping_file') != file_key:
            st.session_state['mapping_done'] = False
            st.session_state['mapping_file'] = file_key

        if not st.session_state.get('mapping_done', False):
            st.info('Mapeie as colunas do arquivo para o formato esperado pelo app:')
            with st.form('form_colunas'):
                # sugerir colunas detectadas como default
                col_date_sel = st.selectbox('Coluna de Data (pode conter data+hora)', df_raw.columns, index=0 if col_date is None else list(df_raw.columns).index(col_date))
                col_local_sel = st.selectbox('Coluna de Local', df_raw.columns, index=0 if col_local is None else list(df_raw.columns).index(col_local))
                col_demanda_sel = st.selectbox('Coluna de Demanda', df_raw.columns, index=0 if col_demanda is None else list(df_raw.columns).index(col_demanda))
                col_hora_sel = None
                if col_hora is not None:
                    col_hora_sel = st.selectbox('Coluna de Hora (opcional)', [None] + list(df_raw.columns), index=0)

                # Opção: permitir especificar formato de data customizado (strftime)
                date_format_input = st.text_input('Formato de data (opcional, strftime)', value='')

                submitted = st.form_submit_button('Confirmar mapeamento')

            if not submitted:
                st.info('Selecione as colunas e clique em Confirmar mapeamento.')
                st.stop()

            # Validar conversão das datas com o formato fornecido (se houver)
            try:
                if date_format_input:
                    test_dates = pd.to_datetime(df_raw[col_date_sel], format=date_format_input, errors='coerce')
                else:
                    test_dates = pd.to_datetime(df_raw[col_date_sel], errors='coerce', infer_datetime_format=True)

                n_invalid = test_dates.isna().sum()
                if n_invalid > 0:
                    pct = n_invalid / max(1, len(test_dates))
                    st.warning(f'Conversao: {n_invalid} valores invalidos detectados na coluna de data ({pct:.1%}). Ajuste o formato ou selecione a coluna correta.')
                    # Allow proceed, but require confirmation
                    if not st.confirm('Deseja prosseguir mesmo assim (valores invalidos serão filtrados)?'):
                        st.stop()
            except Exception as e:
                st.error(f'Erro ao verificar formato da data: {e}')
                st.stop()

            rename_map = {col_date_sel: 'data', col_local_sel: 'local', col_demanda_sel: 'demanda'}
            if col_hora_sel:
                rename_map[col_hora_sel] = 'hora'

            st.session_state['rename_map'] = rename_map
            st.session_state['date_format_input'] = date_format_input
            st.session_state['mapping_done'] = True
            st.success('Mapeamento confirmado — continue nas outras opções.')
            # permitir refazer mapeamento depois caso necessário
            if st.button('Refazer mapeamento'):
                st.session_state['mapping_done'] = False
                st.experimental_rerun()
        else:
            rename_map = {}
            if col_date:
                rename_map[col_date] = 'data'
            if col_local:
                rename_map[col_local] = 'local'
            if col_demanda:
                rename_map[col_demanda] = 'demanda'
            if col_hora:
                rename_map[col_hora] = 'hora'

        # Use mapping from session_state se existir (persistente após mapeamento)
        if 'rename_map' in st.session_state:
            df = df_raw.rename(columns=st.session_state['rename_map'])
        else:
            df = df_raw.rename(columns=rename_map)
        if 'sub_local' not in df.columns:
            df['sub_local'] = 'Todos'
        if 'local_terciario' not in df.columns:
            df['local_terciario'] = 'Todos'
        # Aplicar conversao de data usando formato customizado do mapeamento (se existir)
        date_fmt = st.session_state.get('date_format_input', '')
        if date_fmt:
            try:
                df['data'] = pd.to_datetime(df['data'], format=date_fmt, errors='coerce')
            except Exception:
                df['data'] = pd.to_datetime(df['data'], errors='coerce', infer_datetime_format=True)
        else:
            df['data'] = pd.to_datetime(df['data'], errors='coerce', infer_datetime_format=True)
        valido, resultado = validar_dados(df)
        if not valido:
            st.error(f"[ERRO] {resultado}")
            st.stop()
        df = resultado
        st.success("[OK] Dados carregados!")

        # Fluxo principal do app (mantido após validação)
        # Calcula período com formatação DD/MM/AA HH:MM para exibição reduzida
        start_dt = df['data'].min()
        end_dt = df['data'].max()
        try:
            start_str = pd.to_datetime(start_dt).strftime('%d/%m/%y %H:%M')
        except Exception:
            start_str = str(start_dt)
        try:
            end_str = pd.to_datetime(end_dt).strftime('%d/%m/%y %H:%M')
        except Exception:
            end_str = str(end_dt)

        col1, col2, col3 = st.columns(3)
        col1.metric("Registros", len(df))
        # Exibir período em fonte menor (menor que a métrica de padrão)
        with col2:
            st.markdown("**Período**")
            st.markdown(f"<div style='font-size:18px'>{start_str} a {end_str}</div>", unsafe_allow_html=True)
        col3.metric("Localizacoes", len(df.groupby(['local', 'sub_local', 'local_terciario'])))

        st.subheader("Selecione Localizacao")
        col1, col2, col3 = st.columns(3)

        locais = sorted(df['local'].dropna().unique())
        locais_opts = ['Todos'] + locais
        # Seleção do nível principal (Local)
        with col1:
            local = st.selectbox('Local', locais_opts, index=0, key='local')
        # (local, sub_local selectbox code continues below)
        if local is not None and str(local).lower() != 'todos':
            sub_locais = sorted(df[df['local'] == local]['sub_local'].dropna().unique())
        else:
            sub_locais = sorted(df['sub_local'].dropna().unique())
        sub_local_opts = ['Todos'] + sub_locais
        # preparar strings para o período no formato DD/MM/AA HH:MM
        start_dt = df['data'].min()
        end_dt = df['data'].max()
        try:
            start_str = pd.to_datetime(start_dt).strftime('%d/%m/%y %H:%M')
        except Exception:
            start_str = str(start_dt)
        try:
            end_str = pd.to_datetime(end_dt).strftime('%d/%m/%y %H:%M')
        except Exception:
            end_str = str(end_dt)

        with col2:
            sub_local = st.selectbox("Nivel1", sub_local_opts, index=0, key='sl')

        # Calcula lista de terciarios depois da escolha de sub_local
        if (local is not None and str(local).lower() != 'todos') and (sub_local is not None and str(sub_local).lower() != 'todos'):
            terciarios = sorted(df[(df['local'] == local) & (df['sub_local'] == sub_local)]['local_terciario'].dropna().unique())
        elif (local is not None and str(local).lower() != 'todos') and (sub_local is not None and str(sub_local).lower() == 'todos'):
            terciarios = sorted(df[df['local'] == local]['local_terciario'].dropna().unique())
        elif (sub_local is not None and str(sub_local).lower() != 'todos'):
            terciarios = sorted(df[df['sub_local'] == sub_local]['local_terciario'].dropna().unique())
        else:
            terciarios = sorted(df['local_terciario'].dropna().unique())
        terciario_opts = ['Todos'] + terciarios

        # Terciario (Nivel2) na terceira coluna
        with col3:
            local_terciario = st.selectbox('Nivel2', terciario_opts, index=0, key='lt')

        if (local is not None and str(local).lower() != 'todos') and (sub_local is not None and str(sub_local).lower() != 'todos'):
            terciarios = sorted(df[(df['local'] == local) & (df['sub_local'] == sub_local)]['local_terciario'].dropna().unique())
        elif (local is not None and str(local).lower() != 'todos') and (sub_local is not None and str(sub_local).lower() == 'todos'):
            terciarios = sorted(df[df['local'] == local]['local_terciario'].dropna().unique())
        elif (sub_local is not None and str(sub_local).lower() != 'todos'):
            terciarios = sorted(df[df['sub_local'] == sub_local]['local_terciario'].dropna().unique())
        else:
            terciarios = sorted(df['local_terciario'].dropna().unique())
        terciario_opts = ['Todos'] + terciarios
        # (Período já exibido no topo em fonte reduzida)
        dados_sel = preparar_dados(df, local, sub_local, local_terciario)

        # Se houver coluna 'hora' ou o campo 'data' contiver horas, oferecer granularidade horária
        tem_hora = 'hora' in df.columns or (pd.api.types.is_datetime64_any_dtype(df['data']) and df['data'].dt.hour.nunique() > 1)

        # Se detectamos hora no timestamp (coluna 'data' com time), sugerir Horária por padrão,
        # mas permitir o usuário forçar agregação diária (override)
        if tem_hora:
            # Se a coluna 'data' já contém horas, setar Horária por padrão
            default_index = 1
            force_daily = st.checkbox('Forçar agregação Diária (ignorar hora no DataHora)', value=False, key='force_daily')
            default_idx = 0 if force_daily else default_index
        else:
            default_idx = 0

        # Usar um único widget com chave para permitir mudança programática via session_state
        granularidade = st.radio("Granularidade", ['Diária', 'Horária', 'Turnos'], index=default_idx, key='granularidade')
        # fallback não necessário

        if granularidade == 'Diária' and tem_hora:
            # opção de agregação ao agrupar por dia
            agg_metodo = st.selectbox('Agregação para diário', ['Soma', 'Média', 'Máximo', 'Mínimo', 'Percentil 80%'], index=0)
            # Normalizar para date se tivermos componente horária (para agregar por dia)
            df_for_agg = df.copy()
            df_for_agg['data'] = pd.to_datetime(df_for_agg['data']).dt.normalize()
            if agg_metodo == 'Soma':
                dados_sel = df_for_agg.groupby(['data', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].sum()
            elif agg_metodo == 'Média':
                dados_sel = df.groupby(['data', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].mean()
            elif agg_metodo == 'Máximo':
                dados_sel = df.groupby(['data', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].max()
            elif agg_metodo == 'Mínimo':
                dados_sel = df.groupby(['data', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].min()
            elif agg_metodo == 'Percentil 80%':
                grouped = df_for_agg.groupby(['data', 'local', 'sub_local', 'local_terciario'])
                dados_sel = grouped['demanda'].quantile(0.8).reset_index()
                dados_sel.rename(columns={'demanda': 'demanda'}, inplace=True)
            else:
                dados_sel = df.groupby(['data', 'local', 'sub_local', 'local_terciario'], as_index=False)['demanda'].max()
            # revalidar dataframe
            valido, resultado = validar_dados(dados_sel)
            if not valido:
                st.error(f"[ERRO] {resultado}")
                st.stop()
            dados_sel = resultado

            # Não oferecer agregação por turno dentro da granularidade 'Diária':
            # Se o usuário quiser trabalhar com turnos, deve selecionar a granularidade 'Turnos'.
            if tem_hora:
                st.info('Você selecionou Granularidade = Diária. Para agregar por turnos, escolha Granularidade = Turnos.')
                # Botão 'Ir para Turnos' removido — usuário deve selecionar Turnos manualmente

        if granularidade == 'Horária':
            # construir timestamp horário a partir de data + hora caso exista coluna 'hora'
            if 'hora' in df.columns:
                df['data'] = pd.to_datetime(df['data']) + pd.to_timedelta(df['hora'], unit='h')
                # re-filtrar por local
                dados_sel = preparar_dados(df, local, sub_local, local_terciario)
            else:
                # data já contém hora
                df['data'] = pd.to_datetime(df['data'])
                dados_sel = preparar_dados(df, local, sub_local, local_terciario)

            # Opções de horários de funcionamento (aparece sempre quando granularidade for Horária)
            with st.expander('Horários de Funcionamento (apenas para previsões)', expanded=False):
                start_hour = st.slider('Hora início', 0, 23, 11, key='start_hour')
                end_hour = st.slider('Hora fim', 0, 23, 22, key='end_hour')
                aplicar_restricao_treino = st.checkbox('Aplicar restrição de funcionamento também ao TREINO (opcional)', value=False, key='aplicar_restricao_treino')
                if aplicar_restricao_treino:
                    st.info('Ao aplicar restrição no TREINO, o modelo usará somente os horários selecionados; previsões serão geradas apenas para esses horários.')
                else:
                    st.info('Sem restrição no TREINO: o modelo treinará em todos os horários; as previsões exibidas podem ser filtradas para os horários de funcionamento.')

        # Turnos: configuração
        if granularidade == 'Turnos':
            if not tem_hora:
                st.warning('Turnos requerem dados com componente horário. Ative a opção Horária ou converta seus dados para horário.')
                granularidade = 'Diária'
            else:
                with st.expander('Configurar Turnos (agregação de horas)', expanded=True):
                    n_turnos = st.number_input('Número de turnos por dia (máx. 3)', min_value=1, max_value=3, value=2)
                    turnos = []
                    for i in range(int(n_turnos)):
                        c1, c2 = st.columns(2)
                        with c1:
                            s = st.slider(f'Turno {i+1} - Hora início', 0, 23, 11 + i*6, key=f'start_{i}')
                        with c2:
                            e = st.slider(f'Turno {i+1} - Hora fim', 0, 23, 16 + i*6, key=f'end_{i}')
                        turnos.append((int(s), int(e)))
                    agg_turno = st.selectbox('Agregação por turno (apenas para dados horários)', ['Soma', 'Média', 'Máximo', 'Mínimo', 'Percentil 80%'], index=0)
                    aplicar_restricao_treino = st.checkbox('Aplicar restrição de treino ao conjunto agregado por turnos', value=True, key='aplicar_restricao_treino_turno')
                # Agregar dados por turno
                try:
                    # Construir dados horários se necessário
                    df_para_turno = df.copy()
                    # Forcar conversao segura para datetimes — use errors='coerce' para detectar valores invalidos
                    if 'hora' in df_para_turno.columns:
                        df_para_turno['data'] = pd.to_datetime(df_para_turno['data'], errors='coerce') + pd.to_timedelta(df_para_turno['hora'], unit='h')
                    else:
                        df_para_turno['data'] = pd.to_datetime(df_para_turno['data'], errors='coerce')

                    # Mostrar linhas com timestamps invalidos (se existirem)
                    n_invalid = df_para_turno['data'].isna().sum()
                    if n_invalid > 0:
                        st.warning(f"Encontradas {n_invalid} linhas com timestamp invalido em 'data' — verifique formato (ex.: 'YYYY-MM-DD HH:MM' ou coluna 'hora').")
                    
                    # Verificar se realmente temos dados com horas diferentes
                    horas_unicas = df_para_turno['data'].dt.hour.nunique()
                    if horas_unicas <= 1:
                        st.warning('Dataset não contém dados em múltiplos horários. Usando agregação Diária.')
                        dados_sel = preparar_dados(df, local, sub_local, local_terciario)
                        if 'hora' in dados_sel.columns:
                            dados_sel['data'] = pd.to_datetime(dados_sel['data'])
                    else:
                        # Provide mapping preview so user can confirm which rows fall into each turno
                        with st.expander('Confirmar mapeamento de turnos (exibir amostra)', expanded=False):
                            # If the 'data' column has time component, show it; otherwise warn
                            if 'data' in df_para_turno.columns and pd.api.types.is_datetime64_any_dtype(df_para_turno['data']):
                                st.info('A coluna `data` contém componente horário — será utilizada como timestamp (Data + Hora).')
                            else:
                                st.warning('A coluna `data` não contém componente horário — verifique os dados ou adicione a coluna `hora`.')

                            # Compute mapping sample
                            st.markdown("""
                            **Como funcionam turnos que cruzam a meia-noite (wrap-around)**

                            - Se o `start` do turno for maior que o `end` (ex.: `(22, 5)`), significa que o turno começa no dia atual e termina no dia seguinte.
                            - Horários >= `start` (ex.: 22:00, 23:00) pertencem ao turno do dia atual.
                            - Horários <= `end` (ex.: 00:00, 02:00) pertencem ao turno do mesmo `turno` porém o `shift_date` será o dia anterior, porque o turno começou na noite anterior.
                            - Exemplo: com `(22, 5)`, `2025-11-01 23:00` -> `Turno 1` com `shift_date` = `2025-11-01`; `2025-11-02 02:00` -> `Turno 1` com `shift_date` = `2025-11-01`.
                            """)
                            try:
                                df_map = df_para_turno.copy()
                                df_map['turno_idx'] = df_map['data'].apply(lambda d: _assign_shift(d, turnos))
                                df_map['turno'] = df_map['turno_idx'].apply(lambda x: f"Turno {int(x)+1}" if not pd.isna(x) else None)

                                def _compute_shift_day_row(r):
                                    idx = r['turno_idx']
                                    if pd.isna(idx):
                                        return None
                                    idx = int(idx)
                                    s, e = turnos[idx]
                                    dt = r['data']
                                    hr = dt.hour
                                    if s <= e:
                                        return pd.Timestamp(dt.date())
                                    else:
                                        if hr <= e:
                                            return pd.Timestamp((dt - pd.Timedelta(days=1)).date())
                                        return pd.Timestamp(dt.date())

                                df_map['shift_date'] = df_map.apply(_compute_shift_day_row, axis=1)
                                # show sample rows
                                # pick a sensible demand column if exists
                                demand_col = next((c for c in df_map.columns if c.lower() in ['demanda', 'mesas ocupadas', 'mesas_ocupadas', 'count', 'qtd', 'quant', 'value']), None)
                                cols_to_show = ['data', 'turno_idx', 'turno', 'shift_date']
                                if demand_col:
                                    cols_to_show.insert(1, demand_col)
                                st.dataframe(df_map[cols_to_show].head(50))
                                st.write('Contagem por turno (inclui None para não mapeados):')
                                st.table(df_map['turno'].value_counts(dropna=False).rename_axis('Turno').reset_index(name='Contagem'))
                                # Exemplo wrap-around: amostra demonstração
                                if st.button('Exemplo wrap-around', key='ex_wrap'):
                                    # Escolhe turno demonstrativo (faz exemplo com wrap-around se tiver um turno assim)
                                    sample_shifts = turnos if any(s > e for s, e in turnos) else [(22, 5), (6, 13)]
                                    # Criar amostra de horários representativos nos dois dias
                                    base_day = pd.Timestamp(df_map['data'].dt.date.min())
                                    samples = [
                                        pd.Timestamp(base_day) + pd.Timedelta(hours=h)
                                        for h in [21, 22, 23, 0, 1, 2, 6, 12, 14]
                                    ]
                                    df_sample = pd.DataFrame({'data': samples})
                                    df_sample['turno_idx'] = df_sample['data'].apply(lambda d: _assign_shift(d, sample_shifts))
                                    df_sample['turno'] = df_sample['turno_idx'].apply(lambda x: f"Turno {int(x)+1}" if not pd.isna(x) else None)

                                    def _compute_shift_day_demo(r):
                                        idx = r['turno_idx']
                                        if pd.isna(idx):
                                            return None
                                        idx = int(idx)
                                        s, e = sample_shifts[idx]
                                        dt = r['data']
                                        hr = dt.hour
                                        if s <= e:
                                            return pd.Timestamp(dt.date())
                                        else:
                                            if hr <= e:
                                                return pd.Timestamp((dt - pd.Timedelta(days=1)).date())
                                            return pd.Timestamp(dt.date())

                                    df_sample['shift_date'] = df_sample.apply(_compute_shift_day_demo, axis=1)
                                    st.write('Exemplo (wrap-around) — usando turnos: ' + str(sample_shifts))
                                    st.dataframe(df_sample)
                            except Exception as e:
                                logging.exception(f"Erro ao calcular preview de mapeamento de turnos: {e}")
                                st.error(f"Erro ao calcular preview de mapeamento: {e}")

                        df_turnos = aggregate_shifts(df_para_turno, turnos, agg=agg_turno)
                        if df_turnos is None:
                            st.error('Erro ao agregar turnos. Verifique os dados e tente novamente.')
                            st.stop()
                        # revalidar
                        valido, resultado = validar_dados(df_turnos)
                        if not valido:
                            st.error(f"[ERRO] {resultado}")
                            st.stop()
                        dados_sel = resultado
                        # garantir coluna 'turno' presente na saida (padrão necessario para visualizacao/analise)
                        if 'turno' not in dados_sel.columns:
                            # se existir turno_idx, criar label 'Turno N', caso contrario criar branco
                            if 'turno_idx' in dados_sel.columns:
                                dados_sel['turno'] = dados_sel['turno_idx'].apply(lambda x: f"Turno {int(x)+1}" if not pd.isna(x) else None)
                            else:
                                dados_sel['turno'] = None
                        # salvar turnos em session_state para uso posterior (e.g., gerar datas futuras)
                        try:
                            st.session_state['turnos'] = turnos
                            st.success('Turnos salvos na sessão.')
                        except Exception:
                            logging.info('Não foi possível gravar turnos em session_state; prosseguindo sem persistência.')
                        # (não precisa de exceção adicional aqui)

                        # (Trend plot by turno removed per user request)
                        except Exception:
                            # não interromper se não for possível gravar em session_state (pode ocorrer após criação de widget)
                            logging.info('Não foi possível gravar turnos em session_state; prosseguindo sem persistência.')
                except Exception as e:
                    logging.exception(f"Erro agregando turnos: {e}")
                    st.error('Erro ao agregar turnos, ver logs.')

        if len(dados_sel) < 10:
            st.warning("[AVISO] Dados insuficientes (min 10).")
        else:
            # Sempre exibir o horizonte em dias; ao projetar horário multiplicamos por 24
            horizon_days = st.slider("Horizonte (dias)", 1, 90, 7)
            if granularidade == 'Horária':
                # converter para horas. Se o treino for restrito aos horários de funcionamento, usa "horas de funcionamento por dia"
                if aplicar_restricao_treino:
                    open_hours_per_day = max(0, (end_hour - start_hour) + 1)
                    open_hours_per_day = max(2, open_hours_per_day)
                    horizonte = int(horizon_days * open_hours_per_day)
                else:
                    horizonte = int(horizon_days * 24)
            elif granularidade == 'Turnos':
                # Para turnos, o slider continua representando dias corridos —
                # cada dia possui N turnos, então multiplicamos por N para obter
                # o número de passos (previsões) a serem geradas.
                n_turnos = len(st.session_state.get('turnos', [])) or 1
                horizonte = int(horizon_days * n_turnos)
                st.info(f'Horizonte = {horizon_days} dias → {horizonte} previsões (N={n_turnos} turnos/dia)')
            else:
                horizonte = int(horizon_days)
            st.subheader("Serie Temporal")
            if not PLOTLY_AVAILABLE:
                st.error("Plotly não está instalado no ambiente Python. Execute `pip install plotly` no venv ou use: & \"<venv_path>\" -m pip install plotly")
                st.stop()
            fig_orig = go.Figure(data=[go.Scatter(x=dados_sel['data'], y=dados_sel['demanda'], mode='lines+markers')])
            fig_orig.update_layout(title=f"Demanda - {local} / Nivel1: {sub_local} / Nivel2: {local_terciario}", height=400, xaxis_title="Data", yaxis_title="Demanda")
            st.plotly_chart(fig_orig, use_container_width=True, config={'displayModeBar': False, 'responsive': True})

            # Drilldown removido — comportamento original sem drilldown

            if st.button("Gerar Previsoes", key='exec'):
                prog = st.progress(0)
                status = st.empty()
                resultados = {}
                # Ajustes de sazonalidade conforme granularidade
                if granularidade == 'Horária':
                    if aplicar_restricao_treino:
                        orig_open_hours = max(0, (end_hour - start_hour) + 1)
                        if orig_open_hours < 2:
                            st.warning('A sazonalidade exige periodicidade > 1; o app usará 2 como valor mínimo para sazonalidade.')
                        seasonal_s = max(2, orig_open_hours)
                    else:
                        seasonal_s = 24
                    freq = 'H'
                elif granularidade == 'Turnos':
                    # Cada dia tem N turnos definidos
                    n_turnos = len(st.session_state.get('turnos', [])) or 1
                    if n_turnos < 2:
                        st.warning('Definido 1 turno por dia — alguns modelos requerem periodicidade > 1; usando 2 como valor mínimo.')
                    seasonal_s = max(2, n_turnos)
                    freq = 'H'
                else:
                    seasonal_s = 7
                    freq = 'D'

                # Definir modelos — Prophet receberá future_custom quando usamos Turnos
                modelos = None

                # Gerar datas futuras para exibição e (quando aplicável) para passar ao Prophet
                if granularidade == 'Horária':
                    if aplicar_restricao_treino:
                        def gerar_futuro_horas_abertura(ultimo_ts, passos, sh, eh):
                            res = []
                            current = ultimo_ts + timedelta(hours=1)
                            while len(res) < passos:
                                day = current.date()
                                for h in range(sh, eh + 1):
                                    cand = pd.Timestamp(day) + pd.Timedelta(hours=h)
                                    if cand >= current:
                                        res.append(cand)
                                        if len(res) >= passos:
                                            break
                                current = pd.Timestamp(day + pd.Timedelta(days=1)) + pd.Timedelta(hours=sh)
                            return pd.DatetimeIndex(res)

                        ultimo = dados_sel['data'].max()
                        datas_fut = gerar_futuro_horas_abertura(ultimo, horizonte, start_hour, end_hour)
                    else:
                        datas_fut = pd.date_range(start=dados_sel['data'].max() + timedelta(hours=1), periods=horizonte, freq='H')
                elif granularidade == 'Turnos':
                    def gerar_futuro_turnos(ultimo_ts, passos, shift_starts):
                        res = []
                        current = ultimo_ts + timedelta(hours=1)
                        min_start = min(shift_starts)
                        while len(res) < passos:
                            day = current.date()
                            for s in shift_starts:
                                cand = pd.Timestamp(day) + pd.Timedelta(hours=s)
                                if cand >= current:
                                    res.append(cand)
                                    if len(res) >= passos:
                                        break
                            current = pd.Timestamp(day + pd.Timedelta(days=1)) + pd.Timedelta(hours=min_start)
                        return pd.DatetimeIndex(res)

                    ultimo = dados_sel['data'].max()
                    turnos_starts = [s for s, e in st.session_state.get('turnos', [(11,22)])]
                    datas_fut = gerar_futuro_turnos(ultimo, horizonte, turnos_starts)
                else:
                    datas_fut = pd.date_range(start=dados_sel['data'].max() + timedelta(days=1), periods=horizonte)
                metricas = {}
                # Avaliar modelos (criar lambda para Prophet se turnos for selecionado)
                if granularidade == 'Turnos':
                    prophet_lambda = lambda df, h, fut=datas_fut: prever_prophet(df, h, freq=freq, future_custom=fut)
                else:
                    prophet_lambda = lambda df, h: prever_prophet(df, h, freq=freq)

                modelos = [
                    ('Prophet', prophet_lambda),
                    ('ARIMA', lambda df, h: prever_arima(df, h)),
                    ('SARIMA', lambda df, h: prever_sarima(df, h, seasonal_periods=seasonal_s)),
                    ('Holt-Winters', lambda df, h: prever_holt_winters(df, h, seasonal_periods=seasonal_s))
                ]

                for idx, (nome, func) in enumerate(modelos):
                    status.text(f"Processando {nome}...")
                    # Se aplicacao de restricao ao treino for ativada, filtrar dados antes do treino
                    if granularidade == 'Horária' and aplicar_restricao_treino:
                        dados_treino = dados_sel[dados_sel['data'].dt.hour.between(start_hour, end_hour)].copy()
                    elif granularidade == 'Turnos':
                        # dados_sel ja foi agregado por turno; treinar com todos os turnos
                        dados_treino = dados_sel
                    else:
                        dados_treino = dados_sel

                    # Garantir transformacoes e tipos corretos antes do treino
                    valido_treino, dados_treino_validos = validar_dados(dados_treino)
                    if not valido_treino:
                        st.warning(f"Dados insuficientes ou invalidos para treinar {nome}: {dados_treino_validos}")
                        resultados[nome] = None
                        continue

                    previ, erro = func(dados_treino_validos, horizonte)
                    if erro:
                        st.warning(f"{nome}: {erro}")
                        resultados[nome] = None
                        logging.info(f"{nome} erro: {erro}")
                    else:
                        resultados[nome] = previ
                        # calcular horizonte de validacao coerente com a granularidade (em passos)
                        # validar com horizonte coerente: se treino restringido, numero de passos por dia eh o numero de horas de abertura
                        if granularidade == 'Horária' and aplicar_restricao_treino:
                            open_hours = max(1, (end_hour - start_hour) + 1)
                            val_h = int(horizon_days * open_hours)
                        elif granularidade == 'Horária':
                            val_h = int(horizon_days * 24)
                        elif granularidade == 'Turnos':
                            n_turnos = len(st.session_state.get('turnos', [])) or 1
                            val_h = int(horizon_days * n_turnos)
                        else:
                            val_h = horizon_days
                        met = validar_modelo_com_holdout(dados_treino, func, horizonte=val_h, test_size=0.2)
                        metricas[nome] = met
                    prog.progress((idx + 1) / len(modelos))
                status.text("Concluido")
                # Informar ao usuario que limites inferiores foram ajustados para nao-negatividade
                st.info('Observação: limites inferiores (IC) foram ajustados para não serem negativos (valores < 0 são convertidos para 0).')
                tabs = st.tabs(['Comparacao', 'Metricas', 'Prophet', 'ARIMA', 'SARIMA', 'Holt-Winters'])
                # datas_fut ja foi gerado antes da execucao dos modelos
                    
                with tabs[0]:
                        st.subheader("Tabela Consolidada")
                        tab_cons = pd.DataFrame({'data': datas_fut})
                        
                        for nome, previ in resultados.items():
                            if previ is not None:
                                col_p = 'yhat' if 'yhat' in previ.columns else 'previsao'
                                # Garantir que os valores sejam numericos (coerce a NaN caso contrario)
                                try:
                                    tab_cons[nome] = pd.to_numeric(previ[col_p], errors='coerce').values
                                except Exception:
                                    tab_cons[nome] = np.nan
                        
                        # Se a granularidade for 'Turnos', adicionar coluna 'turno' ao resultado
                        if granularidade == 'Turnos':
                            shifts_for_mapping = st.session_state.get('turnos', None)
                            if shifts_for_mapping is None and 'turnos' in locals():
                                shifts_for_mapping = turnos
                            if shifts_for_mapping is None:
                                shifts_for_mapping = [(11, 22)]

                            def map_to_turno(ts):
                                idx = _assign_shift(pd.to_datetime(ts), shifts_for_mapping)
                                # _assign_shift pode retornar None; usar pd.isna para cobrir NaN/None
                                if pd.isna(idx):
                                    return None
                                try:
                                    return f"Turno {int(idx)+1}"
                                except Exception:
                                    return None

                            tab_cons['turno'] = tab_cons['data'].apply(map_to_turno)

                        cols_p = [c for c in tab_cons.columns if c != 'data']
                        # Excluir colunas não numéricas (e.g., 'turno') antes de calcular a média
                        numeric_cols = [c for c in cols_p if c != 'turno']
                        if numeric_cols:
                            try:
                                # forçar conversão para numérico; valores não convertíveis viram NaN e
                                # não interrompem o cálculo da média
                                tab_cons[numeric_cols] = tab_cons[numeric_cols].apply(pd.to_numeric, errors='coerce')
                                tab_cons['Media'] = tab_cons[numeric_cols].mean(axis=1)
                            except Exception as e:
                                logging.exception(f"Erro calculando media na tabela consolidada: {e}")
                                tab_cons['Media'] = np.nan
                        
                        # Salvar resultados, metricas, tabela consolidada e dados selecionados no session_state
                        try:
                            st.session_state['resultados'] = resultados
                            st.session_state['metricas'] = metricas
                            st.session_state['tab_cons'] = tab_cons
                            st.session_state['dados_sel'] = dados_sel
                            st.session_state['granularidade'] = granularidade
                            st.session_state['horizon_days'] = horizon_days
                            st.session_state['start_hour'] = start_hour if 'start_hour' in locals() else None
                            st.session_state['end_hour'] = end_hour if 'end_hour' in locals() else None
                            st.session_state['aplicar_restricao_treino'] = aplicar_restricao_treino if 'aplicar_restricao_treino' in locals() else False
                        except Exception:
                            logging.info('Nao foi possivel gravar em session_state')

                        # Se não aplicamos restricao de treino mas queremos exibir somente horarios de funcionamento, filtrar tabela consolidada
                        if granularidade == 'Horária' and not aplicar_restricao_treino:
                            if 'data' in tab_cons.columns:
                                tab_cons = tab_cons[tab_cons['data'].dt.hour.between(start_hour, end_hour)].reset_index(drop=True)
                        st.dataframe(tab_cons.round(2), width='content', height=600)
                        
                        # Botoes de download
                        col1, col2 = st.columns(2)
                        with col1:
                            csv = tab_cons.to_csv(index=False).encode('utf-8')
                            st.download_button('Download CSV', data=csv, file_name=f'previ_{local}_{sub_local}.csv', mime='text/csv')
                        
                        with col2:
                            # Preparar metricas para Excel
                            metricas_list = []
                            for modelo_nome, met in metricas.items():
                                if met is not None:
                                    metricas_list.append({
                                        'Modelo': modelo_nome,
                                        'MAE': met.get('MAE', np.nan),
                                        'RMSE': met.get('RMSE', np.nan),
                                        'MAPE (%)': met.get('MAPE', np.nan)
                                    })
                            
                            df_metricas_export = pd.DataFrame(metricas_list)
                            excel_buffer = criar_excel_com_resultados(tab_cons, df_metricas_export, local, sub_local, local_terciario)
                            st.download_button(
                                label='Download Excel (com metricas)',
                                data=excel_buffer,
                                file_name=criar_nome_arquivo_excel(local, sub_local),
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
                            try:
                                st.session_state['df_metricas_export'] = df_metricas_export
                            except Exception:
                                pass
                        
                        # Grafico comparativo removido por solicitacao; manter apenas a tabela consolidada e downloads
                    
                with tabs[1]:
                        st.subheader("Metricas de Assertividade dos Modelos")

                        # Recuperar resultados, metricas e dados selecionados gravados em session_state (se existirem)
                        resultados = st.session_state.get('resultados', {})
                        metricas = st.session_state.get('metricas', {})
                        tab_cons = st.session_state.get('tab_cons', None)
                        dados_sel_ss = st.session_state.get('dados_sel', None)

                        # Definir lista de modelos localmente para uso quando o checkbox for ativado
                        if 'granularidade' in locals() and granularidade == 'Horária':
                            # se aplicou restrição ao treino, usar numero de horas de abertura como seasonal_periods
                            aplicar_restricao = st.session_state.get('aplicar_restricao_treino', False)
                            if aplicar_restricao:
                                sh = st.session_state.get('start_hour', 11) or 11
                                eh = st.session_state.get('end_hour', 22) or 22
                                sp = max(1, (eh - sh) + 1)
                            else:
                                sp = 24
                            freq_local = 'H'
                        else:
                            st.warning('Nenhum resultado numérico disponivel para calculo de media — verifique output dos modelos.')
                            sp = 7
                            freq_local = 'D'
                        modelos = [
                            ('Prophet', lambda df, h: prever_prophet(df, h, freq=freq_local)),
                            ('ARIMA', lambda df, h: prever_arima(df, h)),
                            ('SARIMA', lambda df, h: prever_sarima(df, h, seasonal_periods=sp)),
                            ('Holt-Winters', lambda df, h: prever_holt_winters(df, h, seasonal_periods=sp))
                        ]

                        if metricas:
                            # Preparar dados de metricas para exibicao
                            metricas_data = []
                            for modelo_nome, met in metricas.items():
                                if met is not None:
                                    metricas_data.append({
                                        'Modelo': modelo_nome,
                                        'MAE': met.get('MAE', np.nan),
                                        'RMSE': met.get('RMSE', np.nan),
                                        'MAPE (%)': met.get('MAPE', np.nan)
                                    })
                            
                            if metricas_data:
                                df_metricas = pd.DataFrame(metricas_data)
                                df_metricas_sorted = df_metricas.sort_values('RMSE')
                                
                                # Exibir tabela de metricas
                                st.write("**Tabela de Metricas (ordenada por RMSE - menor eh melhor):**")
                                st.dataframe(df_metricas_sorted.round(4), width='content')
                                
                                # Ranking visual
                                st.write("**Ranking de Modelos:**")
                                for ranking, (idx, row) in enumerate(df_metricas_sorted.iterrows(), 1):
                                    badge = "1o" if ranking == 1 else f"{ranking}o"
                                    st.write(f"{badge} lugar: **{row['Modelo']}** - RMSE: {row['RMSE']:.4f}")
                                
                                # Indicador visual do melhor modelo
                                melhor = df_metricas_sorted.iloc[0]
                                st.success(f"Melhor modelo: **{melhor['Modelo']}** com RMSE = {melhor['RMSE']:.4f}")
                                
                                # Ranking por Nivel2 (ex-local terciario) - exibição automática
                                st.write('---')
                                st.subheader('Ranking por Nivel2')
                                if dados_sel_ss is None:
                                    st.info("Execute 'Gerar Previsoes' primeiro para calcular ranking por Nivel2.")
                                else:
                                    terciarios = sorted(dados_sel_ss['local_terciario'].dropna().unique())
                                    if not terciarios:
                                        st.info('Nenhum Nivel2 encontrado nos dados selecionados.')
                                    else:
                                        melhores = []
                                        resultados_nivel2 = {}
                                        # Calcular metricas por Nivel2, coletando resultados (minimizar saída durante o loop)
                                        for tert in terciarios:
                                            rows_tert = []
                                            df_tert = dados_sel_ss[dados_sel_ss['local_terciario'] == tert]
                                            for modelo_nome, func in modelos:
                                                # obter horizon_days da sessao caso nao exista localmente
                                                hd = st.session_state.get('horizon_days', 7)
                                                gd = st.session_state.get('granularidade', 'Diária')
                                                if gd != 'Horária':
                                                    eval_horiz = hd
                                                else:
                                                    aplicar_restr = st.session_state.get('aplicar_restricao_treino', False)
                                                    if aplicar_restr:
                                                        sh = st.session_state.get('start_hour', 11) or 11
                                                        eh = st.session_state.get('end_hour', 22) or 22
                                                        open_hours = max(1, (eh - sh) + 1)
                                                        eval_horiz = int(hd * open_hours)
                                                    else:
                                                        eval_horiz = int(hd * 24)
                                                met_t = validar_modelo_com_holdout(df_tert, func, horizonte=eval_horiz, test_size=0.2)
                                                if met_t is not None:
                                                    rows_tert.append({
                                                        'Modelo': modelo_nome,
                                                        'MAE': met_t.get('MAE', np.nan),
                                                        'RMSE': met_t.get('RMSE', np.nan),
                                                        'MAPE (%)': met_t.get('MAPE', np.nan)
                                                    })
                                            if rows_tert:
                                                df_tert_metricas = pd.DataFrame(rows_tert).sort_values('RMSE')
                                                resultados_nivel2[tert] = df_tert_metricas
                                                best = df_tert_metricas.iloc[0]
                                                melhores.append({'Nivel2': tert, 'MelhorModelo': best['Modelo'], 'RMSE': best['RMSE']})
                                            else:
                                                resultados_nivel2[tert] = None

                                        # Mostrar tabela consolidada dos melhores por Nivel2 (padrão)
                                        if melhores:
                                            df_melhores = pd.DataFrame(melhores).sort_values('RMSE')
                                            st.write('---')
                                            st.subheader('Resumo consolidado: Melhor modelo por Nivel2')
                                            st.dataframe(df_melhores.reset_index(drop=True).round(4), width='content')

                                            # Expanders com detalhes por Nivel2 (colapsados por padrão)
                                            st.write('Detalhes por Nivel2 (expanda para ver as metricas por modelo):')
                                            for tert in df_melhores['Nivel2'].tolist():
                                                df_det = resultados_nivel2.get(tert)
                                                with st.expander(f"{tert}"):
                                                    if df_det is None or df_det.empty:
                                                        st.write('Dados insuficientes para calcular metricas neste Nivel2.')
                                                    else:
                                                        st.table(df_det.round(4))
                                                        best = df_det.iloc[0]
                                                        st.success(f"Melhor modelo para {tert}: **{best['Modelo']}** (RMSE = {best['RMSE']:.4f})")

                                        # Salvar resultados no session_state para possivel reuso
                                        try:
                                            st.session_state['ranking_nivel2'] = resultados_nivel2
                                            st.session_state['melhores_nivel2'] = melhores
                                        except Exception:
                                            logging.info('Nao foi possivel gravar ranking_nivel2 no session_state')
                            else:
                                st.warning("Nenhuma metrica disponivel para os modelos selecionados.")
                        else:
                            st.info("Metricas serão calculadas durante a execução dos modelos.")

                        st.write("---")
                        st.write("**Explicacao das metricas:**")
                        st.write("- **MAE (Mean Absolute Error)**: Erro medio absoluto entre previsoes e valores reais.")
                        st.write("- **RMSE (Root Mean Square Error)**: Raiz do erro quadrado medio; penaliza erros maiores mais fortemente.")
                        st.write("- **MAPE (Mean Absolute Percentage Error)**: Percentual medio de erro; util para comparar modelos em escalas diferentes.")
                    
                for tab_idx, (nome, previ) in enumerate(resultados.items()):
                        with tabs[tab_idx + 2]:
                            st.subheader(f"Modelo {nome}")
                            
                            if previ is None:
                                st.error(f"Erro ao processar {nome}")
                            else:
                                # Escolher coluna de previsao (prophet usa 'yhat', outros usam 'previsao')
                                col_p = 'yhat' if 'yhat' in previ.columns else 'previsao'
                                tab_mod = pd.DataFrame({'data': datas_fut, 'previsao': pd.to_numeric(previ[col_p], errors='coerce').values})

                                # Determinar colunas de intervalo de confianca
                                if 'yhat' in previ.columns:
                                    col_inf = 'yhat_lower'
                                    col_sup = 'yhat_upper'
                                else:
                                    # Para ARIMA/Holt-Winters, o wrapper já retorna 'inf' e 'sup'
                                    col_inf = 'inf' if 'inf' in previ.columns else None
                                    col_sup = 'sup' if 'sup' in previ.columns else None

                                if col_inf and col_inf in previ.columns:
                                    tab_mod['inf'] = pd.to_numeric(previ[col_inf], errors='coerce').values
                                if col_sup and col_sup in previ.columns:
                                    tab_mod['sup'] = pd.to_numeric(previ[col_sup], errors='coerce').values
                                
                                st.dataframe(tab_mod.round(2), width='content', height=600)
                                
                                if not PLOTLY_AVAILABLE:
                                    st.error("Plotly não está instalado no ambiente Python. Instale com `pip install plotly` no seu venv.")
                                    st.stop()
                                # Criar gráfico comparativo: real vs predito
                                # Preparar numero de pontos reais a comparar (ultima janela = horizon_days)
                                steps_per_day = 1
                                if granularidade == 'Horária':
                                    if aplicar_restricao_treino:
                                        steps_per_day = max(1, (end_hour - start_hour) + 1)
                                    else:
                                        steps_per_day = 24
                                elif granularidade == 'Turnos':
                                    steps_per_day = len(st.session_state.get('turnos', [])) or 1

                                n_real_points = int(horizon_days * steps_per_day)

                                # Se não houver pontos suficientes para comparação, mostrar aviso e exibir previsão simples
                                if n_real_points <= 0 or len(dados_sel) <= n_real_points:
                                    st.warning('Não há dados históricos suficientes para exibir comparação real vs predito. Exibindo apenas previsão.')
                                    # Mostrar previsão no formato antigo para análise
                                    fig_mod = go.Figure()
                                    col_p = 'yhat' if 'yhat' in previ.columns else 'previsao'
                                    x_vals = datas_fut
                                    y_vals = previ[col_p].values
                                    fig_mod.add_trace(go.Scatter(x=x_vals, y=y_vals, mode='lines+markers', name='Previsão', line=dict(color='red', dash='dash')))
                                    fig_mod.update_layout(title=f'Previsão - {nome}', height=450, xaxis_title='Data', yaxis_title='Demanda')
                                    st.plotly_chart(fig_mod, use_container_width=True, config={'displayModeBar': False, 'responsive': True})
                                else:
                                    # Treinar modelo sem a janela de holdout para gerar previsões comparáveis
                                    try:
                                        # criar df de treino (todos exceto janela final)
                                        df_train_for_holdout = dados_treino_validos.iloc[:-n_real_points].copy()
                                        # Obter previsoes para a janela final
                                        if nome == 'Prophet':
                                            # Para Prophet, predizer exatamente nas datas do holdout (real_window['data'])
                                            # Gera previsoes alinhadas com timestamps reais
                                            try:
                                                future_idx = pd.DatetimeIndex(real_window['data'])
                                            except Exception:
                                                future_idx = None
                                            previ_hold, err_hold = prever_prophet(df_train_for_holdout, n_real_points, freq=freq, future_custom=future_idx)
                                        elif nome == 'ARIMA':
                                            previ_hold, err_hold = prever_arima(df_train_for_holdout, n_real_points)
                                        elif nome == 'SARIMA':
                                            previ_hold, err_hold = prever_sarima(df_train_for_holdout, n_real_points, seasonal_periods=seasonal_s)
                                        elif nome == 'Holt-Winters':
                                            previ_hold, err_hold = prever_holt_winters(df_train_for_holdout, n_real_points, seasonal_periods=seasonal_s)
                                        else:
                                            previ_hold, err_hold = None, 'Modelo desconhecido'

                                        if err_hold or previ_hold is None:
                                            st.warning(f"{nome}: não foi possível gerar comparação (erro no holdout: {err_hold}). Exibindo previsão padrão.")
                                            fig_mod = go.Figure()
                                            col_p = 'yhat' if 'yhat' in previ.columns else 'previsao'
                                            x_vals = datas_fut
                                            y_vals = previ[col_p].values
                                            fig_mod.add_trace(go.Scatter(x=x_vals, y=y_vals, mode='lines+markers', name='Previsão', line=dict(color='red', dash='dash')))
                                            fig_mod.update_layout(title=f'Previsão - {nome}', height=450, xaxis_title='Data', yaxis_title='Demanda')
                                            st.plotly_chart(fig_mod, use_container_width=True, config={'displayModeBar': False, 'responsive': True})
                                        else:
                                            # Colocar previsao holdout alinhada com datas reais da janela final
                                            real_window = dados_treino_validos.iloc[-n_real_points:].copy()
                                            # Timestamps reais
                                            x_real = real_window['data']
                                            y_real = real_window['demanda']

                                            col_p = 'yhat' if 'yhat' in previ_hold.columns else 'previsao'
                                            y_pred = pd.to_numeric(previ_hold[col_p], errors='coerce').values[:n_real_points]

                                            fig_comp = go.Figure()
                                            fig_comp.add_trace(go.Scatter(x=x_real, y=y_real, mode='lines+markers', name='Real', line=dict(color='blue')))
                                            fig_comp.add_trace(go.Scatter(x=x_real, y=y_pred, mode='lines+markers', name='Previsto', line=dict(color='red', dash='dash')))

                                            # Adicionar IC se disponível (Prophet: yhat_lower/yhat_upper, statsmodels: inf/sup)
                                            if 'yhat_lower' in previ_hold.columns and 'yhat_upper' in previ_hold.columns:
                                                y_inf = pd.to_numeric(previ_hold['yhat_lower'], errors='coerce').values[:n_real_points]
                                                y_sup = pd.to_numeric(previ_hold['yhat_upper'], errors='coerce').values[:n_real_points]
                                                # garantir non-negative e consistencia
                                                y_inf = np.maximum(y_inf, 0)
                                                fig_comp.add_trace(go.Scatter(x=x_real, y=y_inf, fill=None, mode='lines', line=dict(width=0), showlegend=False))
                                                fig_comp.add_trace(go.Scatter(x=x_real, y=y_sup, fill='tonexty', mode='lines', name='IC', line=dict(width=0), fillcolor='rgba(0,100,200,0.2)'))
                                            elif 'inf' in previ_hold.columns and 'sup' in previ_hold.columns:
                                                y_inf = pd.to_numeric(previ_hold['inf'], errors='coerce').values[:n_real_points]
                                                y_sup = pd.to_numeric(previ_hold['sup'], errors='coerce').values[:n_real_points]
                                                y_inf = np.maximum(y_inf, 0)
                                                fig_comp.add_trace(go.Scatter(x=x_real, y=y_inf, fill=None, mode='lines', line=dict(width=0), showlegend=False))
                                                fig_comp.add_trace(go.Scatter(x=x_real, y=y_sup, fill='tonexty', mode='lines', name='IC', line=dict(width=0), fillcolor='rgba(0,100,200,0.2)'))

                                            fig_comp.update_layout(title=f'Real vs Previsto - {nome} (Horizonte: {horizon_days} dias -> {n_real_points} passos)', height=450, xaxis_title='Data', yaxis_title='Demanda')
                                            st.plotly_chart(fig_comp, use_container_width=True, config={'displayModeBar': False, 'responsive': True})
                                    except Exception as e:
                                        logging.exception(f"Erro gerando comparacao real vs predito para {nome}: {e}")
                                        st.error(f"Erro ao gerar comparação para {nome}: {e}")
except Exception as e:
    logging.exception(f"Erro geral do app: {e}")
    st.error('Ocorreu um erro. Consulte logs/app.log para detalhes.')

else:
    st.info('Faca upload CSV para comecar previsoes.')
    # Observação: seção de 'Estrutura Esperada' removida conforme solicitado.
    # Informação sobre dados horários removida por solicitação
